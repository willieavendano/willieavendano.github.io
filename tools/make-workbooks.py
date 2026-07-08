#!/usr/bin/env python3
"""Regenerable Excel workbook builders for course resources.

Run directly to (re)write every workbook this script knows how to build.
Idempotent: re-running overwrites the target file(s) with the same content.

    python3 tools/make-workbooks.py

Requires openpyxl (pip3 install --user openpyxl).
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Shared style constants (matches assets/css/site.css --blue / --paper-deep)
# ---------------------------------------------------------------------------

CUSHMAN_BLUE = "0E406A"
PAPER_DEEP = "F3EFE7"
INK = "141412"
INK_SOFT = "55504A"

HEADER_FILL = PatternFill("solid", fgColor=CUSHMAN_BLUE)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor=PAPER_DEEP)
SECTION_FILL = PatternFill("solid", fgColor=CUSHMAN_BLUE)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, color=CUSHMAN_BLUE, size=16)
BOLD = Font(bold=True, color=INK)
NOTE_FONT = Font(italic=True, color=INK_SOFT)
THIN = Side(style="thin", color="CCCCCC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# ---------------------------------------------------------------------------
# CS Math: Cornerstone I starter workbook
# ---------------------------------------------------------------------------

FOOD_TABLE_ROWS = [
    # Food, Serving, Calories, Protein (g), Carbs (g), Fat (g)
    ("Egg, large", "1 egg", 72, 6.3, 0.4, 4.8),
    ("White rice, cooked", "1 cup", 205, 4.3, 44.5, 0.4),
    ("Chicken breast, grilled", "100 g", 165, 31.0, 0, 3.6),
    ("Black beans, cooked", "1 cup", 227, 15.2, 40.8, 0.9),
    ("Cafecito con azucar (Cuban coffee, sweet)", "1 cup (8 oz)", 40, 0.5, 8.0, 1.0),
    ("Plantain, fried (maduros)", "1 cup", 250, 1.4, 47.0, 8.6),
    ("Ropa vieja (shredded beef)", "1 cup", 285, 26.0, 8.0, 17.0),
    ("Yuca con mojo", "1 cup", 220, 1.6, 52.0, 0.3),
    ("Avocado", "1/2 fruit", 160, 2.0, 8.5, 14.7),
    ("Croqueta (ham)", "1 piece", 130, 4.0, 10.0, 8.0),
    ("Whole wheat bread", "1 slice", 81, 4.0, 13.8, 1.1),
    ("Greek yogurt, plain", "1 cup (170 g)", 100, 17.0, 6.0, 0.7),
    ("Banana", "1 medium", 105, 1.3, 27.0, 0.4),
    ("Pizza slice, cheese", "1 slice", 285, 12.2, 35.7, 10.4),
    ("Cafe con leche", "1 cup (8 oz)", 110, 5.0, 11.0, 5.0),
]

FOOD_TABLE_HEADER_ROW = 1
FOOD_TABLE_FIRST_DATA_ROW = 2
FOOD_TABLE_LAST_DATA_ROW = FOOD_TABLE_FIRST_DATA_ROW + len(FOOD_TABLE_ROWS) - 1  # 16

DAILY_LOG_HEADERS = [
    "Date", "Meal", "Food", "Servings", "Calories", "Protein", "Carbs", "Fat",
]
DAILY_LOG_EXAMPLES = [
    # Date, Meal, Food, Servings — Calories..Fat intentionally left blank
    ("2026-09-21", "Breakfast", "Egg, large", 2),
    ("2026-09-21", "Lunch", "White rice, cooked", 1),
    ("2026-09-21", "Dinner", "Ropa vieja (shredded beef)", 1.5),
]
DAILY_LOG_VALIDATION_ROWS = 200  # ~200 rows of dropdown coverage, per brief


def build_read_me_sheet(wb):
    ws = wb.active
    ws.title = "READ ME FIRST"
    set_widths(ws, [26, 90])

    ws["A1"] = "Cornerstone I -- Food & Macro Tracker"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        (
            "What is a cornerstone?",
            "Every unit in this course ends with a cornerstone project: a working model "
            "you design and build yourself, then present to the class. This workbook is "
            "the starter file for Cornerstone I -- it is NOT the finished project. It's "
            "the frame you build the real thing inside.",
        ),
        (
            "What's already built for you",
            "The Food Table headers and 15 starter foods with real, approximate macro "
            "values (add your own -- aim for 25+ by presentation day). The Daily Log "
            "headers, with a dropdown in the Food column pulled straight from your Food "
            "Table. The Dashboard layout, with labeled boxes waiting for your formulas.",
        ),
        (
            "What you build",
            "1) Grow the Food Table with foods you actually eat. 2) In the Daily Log, "
            "write the VLOOKUP formulas in the Calories/Protein/Carbs/Fat columns that "
            "look up each row's Food in the Food Table and multiply by Servings -- "
            "then keep logging real days. 3) In the Dashboard, write the SUMIF/AVERAGEIF "
            "formulas that summarize your log. 4) Build at least one honest chart from "
            "your own data (no answer charts are pre-built for you).",
        ),
        (
            "No answers here",
            "There are no formulas hidden anywhere in this workbook. Every formula you "
            "see in your finished tracker will be one you wrote. If a cell looks empty, "
            "that's on purpose -- that's your work.",
        ),
        (
            "Presentation date",
            "Thursday, October 15 (Week 8 of the Year Calendar). Bring a working tracker "
            "and one real insight about your own eating patterns -- not just charts, an "
            "actual observation.",
        ),
        (
            "Rubric",
            "Posted in the course Drive folder -- see the README's Google Drive link.",
        ),
    ]

    r = 3
    for heading, body in rows:
        ws.cell(row=r, column=1, value=heading).font = BOLD
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        body_cell = ws.cell(row=r, column=2, value=body)
        body_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 60
        r += 1

    ws.freeze_panes = "A2"


def build_food_table_sheet(wb):
    ws = wb.create_sheet("Food Table")
    headers = ["Food", "Serving", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    set_widths(ws, [34, 16, 12, 13, 12, 10])

    for col, text in enumerate(headers, start=1):
        ws.cell(row=FOOD_TABLE_HEADER_ROW, column=col, value=text)
    style_header_row(ws, FOOD_TABLE_HEADER_ROW, len(headers))

    for i, food_row in enumerate(FOOD_TABLE_ROWS):
        r = FOOD_TABLE_FIRST_DATA_ROW + i
        for col, value in enumerate(food_row, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = BOX
            if col > 1:
                cell.alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                cell.fill = BAND_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A{FOOD_TABLE_HEADER_ROW}:F{FOOD_TABLE_LAST_DATA_ROW}"


def build_daily_log_sheet(wb):
    ws = wb.create_sheet("Daily Log")
    set_widths(ws, [12, 12, 34, 10, 10, 10, 10, 10])

    for col, text in enumerate(DAILY_LOG_HEADERS, start=1):
        ws.cell(row=1, column=col, value=text)
    style_header_row(ws, 1, len(DAILY_LOG_HEADERS))

    for i, example in enumerate(DAILY_LOG_EXAMPLES):
        r = 2 + i
        date_str, meal, food, servings = example
        ws.cell(row=r, column=1, value=date_str)
        ws.cell(row=r, column=2, value=meal)
        ws.cell(row=r, column=3, value=food)
        ws.cell(row=r, column=4, value=servings)
        # Columns 5-8 (Calories..Fat) intentionally left blank -- students build these.
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BOX

    note_row = 6
    ws.merge_cells(f"A{note_row}:H{note_row}")
    note_cell = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Build it: in each Calories / Protein / Carbs / Fat cell above, write a "
            "VLOOKUP that finds this row's Food in the Food Table, then multiply by "
            "Servings. Once that works, keep logging your own real days starting row 7 -- "
            "the Food column dropdown pulls its list from your Food Table automatically."
        ),
    )
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[note_row].height = 45

    # Data validation dropdown on the Food column, sourced from Food Table names.
    # Split around the note row (row 6) so the merged note cell isn't inside the range.
    dv_formula = f"='Food Table'!$A${FOOD_TABLE_FIRST_DATA_ROW}:$A${FOOD_TABLE_LAST_DATA_ROW}"
    dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True, showDropDown=False)
    dv.error = "Choose a food from the Food Table (or add it there first)."
    dv.errorTitle = "Not in Food Table"
    ws.add_data_validation(dv)
    dv.add(f"C2:C4")
    dv.add(f"C7:C{DAILY_LOG_VALIDATION_ROWS + 1}")

    ws.freeze_panes = "A2"


def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")
    set_widths(ws, [24, 16, 16, 16, 44])

    ws["A1"] = "Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    # -- Daily Averages -----------------------------------------------------
    ws.merge_cells("A3:E3")
    ws["A3"] = "Daily Averages"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws["A3"].alignment = Alignment(vertical="center")

    avg_headers = ["Metric", "Value", "", "", "Notes"]
    for col, text in enumerate(avg_headers, start=1):
        if text:
            ws.cell(row=4, column=col, value=text).font = BOLD

    avg_labels = [
        "Average Calories / Day",
        "Average Protein (g) / Day",
        "Average Carbs (g) / Day",
        "Average Fat (g) / Day",
    ]
    for i, label in enumerate(avg_labels):
        r = 5 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2).border = BOX  # empty -- student formula goes here
        note = ws.cell(
            row=r, column=5,
            value="<- your AVERAGEIF formula goes here (Daily Log, matched by date)",
        )
        note.font = NOTE_FONT

    # -- Target vs. Actual ----------------------------------------------------
    section2_row = 10
    ws.merge_cells(f"A{section2_row}:E{section2_row}")
    ws.cell(row=section2_row, column=1, value="Target vs. Actual")
    ws.cell(row=section2_row, column=1).font = SECTION_FONT
    ws.cell(row=section2_row, column=1).fill = SECTION_FILL
    ws.cell(row=section2_row, column=1).alignment = Alignment(vertical="center")

    tva_header_row = section2_row + 1
    for col, text in enumerate(["Metric", "Target", "Actual", "Difference"], start=1):
        ws.cell(row=tva_header_row, column=col, value=text).font = BOLD

    tva_labels = ["Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    tva_first_data_row = tva_header_row + 1
    for i, label in enumerate(tva_labels):
        r = tva_first_data_row + i
        ws.cell(row=r, column=1, value=label)
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BOX  # empty -- student fills in
    tva_last_data_row = tva_first_data_row + len(tva_labels) - 1

    ws.cell(
        row=tva_last_data_row + 1, column=5,
        value="<- Target: your own goal. Actual: pull from Daily Averages above.",
    ).font = NOTE_FONT

    # -- Chart placeholder ------------------------------------------------
    chart_row = tva_last_data_row + 3
    ws.merge_cells(f"A{chart_row}:E{chart_row}")
    ws.cell(row=chart_row, column=1, value="Chart").font = SECTION_FONT
    ws.cell(row=chart_row, column=1).fill = SECTION_FILL
    ws.cell(row=chart_row, column=1).alignment = Alignment(vertical="center")

    note_row = chart_row + 1
    ws.merge_cells(f"A{note_row}:E{note_row + 2}")
    chart_note = ws.cell(
        row=note_row, column=1,
        value=(
            f"Insert your chart here: select A{tva_header_row}:D{tva_last_data_row} "
            "(the Target vs. Actual table), then Insert > Chart > Clustered Column. "
            "Drag the chart into the empty space below this note."
        ),
    )
    chart_note.font = NOTE_FONT
    chart_note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 50


def make_csm_p1_tracker_starter():
    wb = Workbook()
    build_read_me_sheet(wb)
    build_food_table_sheet(wb)
    build_daily_log_sheet(wb)
    build_dashboard_sheet(wb)

    out_path = os.path.join(
        REPO_ROOT, "computer-science-math", "resources", "csm-p1-tracker-starter.xlsx"
    )
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path}")


def main():
    make_csm_p1_tracker_starter()


if __name__ == "__main__":
    main()
