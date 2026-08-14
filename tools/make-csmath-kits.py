#!/usr/bin/env python3
"""Practice and exemplar workbooks for the Computer Science Math kits.

Two workbook families:

  csm-p<N>-*-practice.xlsx   guided drills — the "practice" beat of a kit
  csm-<id>-exemplars.xlsx    the SAME task built four times, one tab per
                             rubric level, plus a "What Changed" tab naming
                             the specific difference between each pair

Run from the repo root:

    python3 tools/make-csmath-kits.py

Idempotent: re-running overwrites the targets with identical content.
Requires openpyxl.
"""

from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from importlib import import_module

_mw = import_module("make-workbooks".replace("-", "_")) if False else None

# ── shared style constants (kept in sync with tools/make-workbooks.py) ──
CUSHMAN_BLUE = "0E406A"
BRIGHT = "0071CE"
PAPER_DEEP = "F3EFE7"
INK_SOFT = "55504A"
GOOD = "2C6E49"
WARN = "8A5A0B"
BAD = "A33A2A"

HEADER_FILL = PatternFill("solid", fgColor=CUSHMAN_BLUE)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor=PAPER_DEEP)
TITLE_FONT = Font(bold=True, color=CUSHMAN_BLUE, size=16)
H2_FONT = Font(bold=True, color=CUSHMAN_BLUE, size=12)
BOLD = Font(bold=True)
NOTE_FONT = Font(italic=True, color=INK_SOFT)
WRAP = Alignment(wrap_text=True, vertical="top")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RES = os.path.join(ROOT, "computer-science-math", "resources")


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def title_block(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    r = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_FONT
        ws["A2"].alignment = WRAP
        r = 3
    return r + 1


def read_me(wb, title, lines, first=True):
    ws = wb.active if first else wb.create_sheet()
    ws.title = "READ ME FIRST"
    set_widths(ws, [30, 96])
    r = title_block(ws, title)
    for label, body in lines:
        ws.cell(row=r, column=1, value=label).font = BOLD
        c = ws.cell(row=r, column=2, value=body)
        c.alignment = WRAP
        ws.row_dimensions[r].height = max(15, 13 * (1 + len(body) // 92))
        r += 1
    return ws


# ═══════════════════════════════════════════════════════════════════
# Cornerstone II practice — HYSA first, then the loan
# ═══════════════════════════════════════════════════════════════════

def build_p2_interest_practice():
    wb = Workbook()
    read_me(wb, "Cornerstone II Practice — Interest, Both Directions", [
        ("What this is",
         "Five drills. Do them in order — each one sets up the next. Nothing here is graded; this is where you "
         "make your mistakes cheaply, before the cornerstone counts."),
        ("Why savings first",
         "Compound interest is one machine. Pointed at your savings it makes you money; pointed at a loan it "
         "costs you money. Same formula, opposite sign. Meet it as a friend first — the amortisation table "
         "stops being frightening once you have already built the growth curve."),
        ("The one rule",
         "Every number that could be computed MUST be computed. If you type a result into a cell, you have "
         "drawn a picture of a model instead of building one. Change an input; if nothing moves, it is broken."),
        ("When you are done",
         "You should be able to answer: why does a longer loan with a smaller monthly payment cost more?"),
    ])

    # ── Drill 1 & 2: HYSA growth ──
    ws = wb.create_sheet("1-2 HYSA Growth")
    set_widths(ws, [26, 15, 15, 15, 15, 46])
    r = title_block(ws, "Drills 1 & 2 — Interest working FOR you",
                    "A high-yield savings account. Build the growth table, then change the compounding.")
    ws.cell(row=r, column=1, value="INPUTS — change these, everything else must follow").font = H2_FONT
    r += 1
    inputs = [("Starting balance", 500, "$#,##0.00"),
              ("Annual rate (APY)", 0.045, "0.00%"),
              ("Compounds per year", 12, "0"),
              ("Years", 5, "0")]
    first_input = r
    for name, val, fmt in inputs:
        ws.cell(row=r, column=1, value=name).font = BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = BAND_FILL
        r += 1
    B_START, B_RATE, B_N, B_YEARS = (f"$B${first_input + i}" for i in range(4))

    r += 1
    ws.cell(row=r, column=1, value="Period").font = HEADER_FONT
    ws.cell(row=r, column=2, value="Balance start").font = HEADER_FONT
    ws.cell(row=r, column=3, value="Interest earned").font = HEADER_FONT
    ws.cell(row=r, column=4, value="Balance end").font = HEADER_FONT
    header_row(ws, r, 4)
    tbl = r + 1
    for i in range(60):
        row = tbl + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2,
                value=f"={B_START}" if i == 0 else f"=D{row-1}").number_format = "$#,##0.00"
        ws.cell(row=row, column=3, value=f"=B{row}*{B_RATE}/{B_N}").number_format = "$#,##0.00"
        ws.cell(row=row, column=4, value=f"=B{row}+C{row}").number_format = "$#,##0.00"
    ws.cell(row=tbl - 1, column=6,
            value="Drill 1: this table is built for you — read every formula and say out loud what it does.").alignment = WRAP
    ws.cell(row=tbl + 1, column=6,
            value="Drill 2: change 'Compounds per year' from 12 to 365. Write one sentence below explaining "
                  "why the ending balance barely moves.").alignment = WRAP
    ws.cell(row=tbl + 4, column=6, value="Your sentence:").font = BOLD
    ws.cell(row=tbl + 5, column=6, value="").fill = BAND_FILL

    chart = LineChart()
    chart.title = "Savings growth"
    chart.y_axis.title = "Balance"
    chart.x_axis.title = "Month"
    chart.height, chart.width = 7, 14
    chart.add_data(Reference(ws, min_col=4, min_row=tbl - 1, max_row=tbl + 59), titles_from_data=True)
    ws.add_chart(chart, f"H{tbl + 8}")

    # ── Drill 3 & 4: the loan ──
    ws = wb.create_sheet("3-4 The Loan")
    set_widths(ws, [26, 15, 15, 15, 15, 46])
    r = title_block(ws, "Drills 3 & 4 — The same machine, pointed at you",
                    "Now you are the one paying. Build the payment, then split it.")
    ws.cell(row=r, column=1, value="INPUTS").font = H2_FONT
    r += 1
    loan_inputs = [("Amount financed", 12000, "$#,##0.00"),
                   ("Annual rate (APR)", 0.07, "0.00%"),
                   ("Term (months)", 60, "0")]
    fi = r
    for name, val, fmt in loan_inputs:
        ws.cell(row=r, column=1, value=name).font = BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = BAND_FILL
        r += 1
    L_AMT, L_APR, L_TERM = (f"$B${fi + i}" for i in range(3))

    ws.cell(row=r, column=1, value="Monthly payment").font = BOLD
    ws.cell(row=r, column=2, value=f"=PMT({L_APR}/12,{L_TERM},-{L_AMT})").number_format = "$#,##0.00"
    ws.cell(row=r, column=6,
            value="Drill 3: PMT does the work. Look it up — what do the three arguments mean? "
                  "Why is the amount negative?").alignment = WRAP
    pay = f"$B${r}"
    r += 2

    for i, h in enumerate(["Month", "Payment", "Interest", "Principal", "Balance"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 5)
    tbl = r + 1
    for i in range(60):
        row = tbl + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=f"={pay}").number_format = "$#,##0.00"
        prev_bal = L_AMT if i == 0 else f"E{row-1}"
        ws.cell(row=row, column=3, value=f"={prev_bal}*{L_APR}/12").number_format = "$#,##0.00"
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "$#,##0.00"
        ws.cell(row=row, column=5, value=f"={prev_bal}-D{row}").number_format = "$#,##0.00"
    ws.cell(row=tbl + 2, column=7,
            value="Drill 4: chart Interest and Principal together. Find the month they cross. "
                  "Explain what the crossover means for someone paying this loan.").alignment = WRAP

    chart = LineChart()
    chart.title = "Interest vs principal"
    chart.y_axis.title = "Dollars"
    chart.x_axis.title = "Month"
    chart.height, chart.width = 7, 14
    chart.add_data(Reference(ws, min_col=3, max_col=4, min_row=tbl - 1, max_row=tbl + 59), titles_from_data=True)
    ws.add_chart(chart, f"G{tbl + 6}")

    # ── Drill 5: the comparison ──
    ws = wb.create_sheet("5 Which Loan")
    set_widths(ws, [30, 18, 18, 18, 52])
    r = title_block(ws, "Drill 5 — The trap",
                    "A smaller monthly payment is not a cheaper loan. Prove it.")
    for i, h in enumerate(["Term", "Monthly payment", "Total paid", "Total interest"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 4)
    r += 1
    for term in (36, 48, 60, 72, 84):
        ws.cell(row=r, column=1, value=f"{term} months")
        ws.cell(row=r, column=2, value=f"=PMT($B$20/12,{term},-$B$19)").number_format = "$#,##0.00"
        ws.cell(row=r, column=3, value=f"=B{r}*{term}").number_format = "$#,##0.00"
        ws.cell(row=r, column=4, value=f"=C{r}-$B$19").number_format = "$#,##0.00"
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Amount financed").font = BOLD
    ws.cell(row=r, column=2, value=12000).number_format = "$#,##0.00"
    ws.cell(row=r, column=2).fill = BAND_FILL
    ws.cell(row=r + 1, column=1, value="Annual rate (APR)").font = BOLD
    ws.cell(row=r + 1, column=2, value=0.07).number_format = "0.00%"
    ws.cell(row=r + 1, column=2).fill = BAND_FILL
    ws.cell(row=r + 3, column=1,
            value="The inputs live in B19 and B20 — the formulas above point at them. "
                  "Change the rate and watch every row move.").alignment = WRAP
    ws.cell(row=r + 5, column=1, value="Which term would you take, and why? (One paragraph.)").font = BOLD

    path = os.path.join(RES, "csm-p2-interest-practice.xlsx")
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════
# Exemplars — the same task, built four ways
# ═══════════════════════════════════════════════════════════════════

EXEMPLAR_TASK_U2 = "Model a $12,000 car loan at 7% APR over 60 months. Report the monthly payment and the total interest paid."

def _ex_header(ws, level, name, gloss, task):
    colour = {4: GOOD, 3: GOOD, 2: WARN, 1: BAD}[level]
    ws["A1"] = f"LEVEL {level} — {name}"
    ws["A1"].font = Font(bold=True, size=15, color=colour)
    ws["A2"] = gloss
    ws["A2"].font = NOTE_FONT
    ws["A3"] = f"Task: {task}"
    ws["A3"].font = NOTE_FONT
    ws["A3"].alignment = WRAP
    return 5


def build_u2_exemplars():
    wb = Workbook()
    read_me(wb, "Cornerstone II — What Each Level Looks Like", [
        ("What this is",
         "The same car-loan task, built four times. Level 4 is what excellence looks like; Level 1 is what "
         "gets handed in when someone runs out of time. Open them side by side."),
        ("How to use it",
         "Read this BEFORE you start building, not after you are graded. Then open 'What Changed' — it names "
         "the exact difference between each pair of levels."),
        ("The fastest way to lose points",
         "Typing in a number that should have been computed. Every level below 3 fails on that alone, no "
         "matter how good it looks."),
        ("Task", EXEMPLAR_TASK_U2),
    ])

    LEVELS = [
        (4, "Advanced", "Publishable. Someone else could pick this up and use it."),
        (3, "Proficient", "Works and is honest. The expected standard."),
        (2, "Developing", "Partly works, or works for the wrong reasons."),
        (1, "Beginning", "Incomplete, or the numbers cannot be trusted."),
    ]

    for level, name, gloss in LEVELS:
        ws = wb.create_sheet(f"Level {level} — {name}")
        set_widths(ws, [28, 16, 15, 15, 15, 50])
        r = _ex_header(ws, level, name, gloss, EXEMPLAR_TASK_U2)

        if level == 4:
            ws.cell(row=r, column=1, value="INPUTS  (change any of these — everything follows)").font = H2_FONT
            r += 1
            base = r
            for nm, val, fmt in [("Vehicle price", 15000, "$#,##0.00"),
                                 ("Down payment", 3000, "$#,##0.00"),
                                 ("Annual rate (APR)", 0.07, "0.00%"),
                                 ("Term (months)", 60, "0")]:
                ws.cell(row=r, column=1, value=nm).font = BOLD
                c = ws.cell(row=r, column=2, value=val); c.number_format = fmt; c.fill = BAND_FILL
                r += 1
            ws.cell(row=r, column=1, value="Amount financed").font = BOLD
            ws.cell(row=r, column=2, value=f"=B{base}-B{base+1}").number_format = "$#,##0.00"
            fin = r; r += 2
            ws.cell(row=r, column=1, value="OUTPUTS").font = H2_FONT
            r += 1
            ws.cell(row=r, column=1, value="Monthly payment").font = BOLD
            ws.cell(row=r, column=2, value=f"=IF(B{base+3}<=0,\"check term\",PMT(B{base+2}/12,B{base+3},-B{fin}))").number_format = "$#,##0.00"
            pay = f"$B${r}"; r += 1
            ws.cell(row=r, column=1, value="Total paid").font = BOLD
            ws.cell(row=r, column=2, value=f"={pay}*B{base+3}").number_format = "$#,##0.00"
            r += 1
            ws.cell(row=r, column=1, value="Total interest").font = BOLD
            ws.cell(row=r, column=2, value=f"=B{r-1}-B{fin}").number_format = "$#,##0.00"
            r += 2
            ws.cell(row=r, column=6, value="Note the IF(): a zero or negative term would produce a nonsense "
                                           "payment, so the model refuses instead of lying.").alignment = WRAP
            for i, h in enumerate(["Month", "Payment", "Interest", "Principal", "Balance"], start=1):
                ws.cell(row=r, column=i, value=h)
            header_row(ws, r, 5)
            t = r + 1
            for i in range(60):
                row = t + i
                ws.cell(row=row, column=1, value=i + 1)
                ws.cell(row=row, column=2, value=f"={pay}").number_format = "$#,##0.00"
                prev = f"$B${fin}" if i == 0 else f"E{row-1}"
                ws.cell(row=row, column=3, value=f"={prev}*$B${base+2}/12").number_format = "$#,##0.00"
                ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "$#,##0.00"
                ws.cell(row=row, column=5, value=f"={prev}-D{row}").number_format = "$#,##0.00"
            ch = LineChart(); ch.title = "Interest vs principal over the life of the loan"
            ch.y_axis.title = "Dollars"; ch.x_axis.title = "Month"; ch.height, ch.width = 7, 15
            ch.add_data(Reference(ws, min_col=3, max_col=4, min_row=r, max_row=t + 59), titles_from_data=True)
            ws.add_chart(ch, f"G{t}")

        elif level == 3:
            ws.cell(row=r, column=1, value="Inputs").font = H2_FONT
            r += 1
            base = r
            for nm, val, fmt in [("Amount financed", 12000, "$#,##0.00"),
                                 ("APR", 0.07, "0.00%"),
                                 ("Term (months)", 60, "0")]:
                ws.cell(row=r, column=1, value=nm).font = BOLD
                c = ws.cell(row=r, column=2, value=val); c.number_format = fmt; c.fill = BAND_FILL
                r += 1
            ws.cell(row=r, column=1, value="Monthly payment").font = BOLD
            ws.cell(row=r, column=2, value=f"=PMT(B{base+1}/12,B{base+2},-B{base})").number_format = "$#,##0.00"
            pay = f"$B${r}"; r += 1
            ws.cell(row=r, column=1, value="Total interest").font = BOLD
            ws.cell(row=r, column=2, value=f"={pay}*B{base+2}-B{base}").number_format = "$#,##0.00"
            r += 2
            for i, h in enumerate(["Month", "Payment", "Interest", "Principal", "Balance"], start=1):
                ws.cell(row=r, column=i, value=h)
            header_row(ws, r, 5)
            t = r + 1
            for i in range(60):
                row = t + i
                ws.cell(row=row, column=1, value=i + 1)
                ws.cell(row=row, column=2, value=f"={pay}").number_format = "$#,##0.00"
                prev = f"$B${base}" if i == 0 else f"E{row-1}"
                ws.cell(row=row, column=3, value=f"={prev}*$B${base+1}/12").number_format = "$#,##0.00"
                ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "$#,##0.00"
                ws.cell(row=row, column=5, value=f"={prev}-D{row}").number_format = "$#,##0.00"
            ch = LineChart(); ch.title = "Interest vs principal"
            ch.height, ch.width = 6, 13
            ch.add_data(Reference(ws, min_col=3, max_col=4, min_row=r, max_row=t + 59), titles_from_data=True)
            ws.add_chart(ch, f"G{t}")
            ws.cell(row=t + 62, column=1,
                    value="Missing vs Level 4: no down-payment input, no edge-case guard, "
                          "axes unlabelled.").alignment = WRAP

        elif level == 2:
            ws.cell(row=r, column=1, value="Car loan").font = H2_FONT
            r += 1
            ws.cell(row=r, column=1, value="Amount"); ws.cell(row=r, column=2, value=12000).number_format = "$#,##0.00"
            r += 1
            ws.cell(row=r, column=1, value="Payment")
            ws.cell(row=r, column=2, value="=PMT(0.07/12,60,-12000)").number_format = "$#,##0.00"
            ws.cell(row=r, column=6, value="⚠ Rate and term are buried INSIDE the formula. Changing the "
                                           "input cell above does nothing.").alignment = WRAP
            ws.cell(row=r, column=6).font = Font(color=WARN)
            pay_row = r
            r += 1
            ws.cell(row=r, column=1, value="Total interest")
            ws.cell(row=r, column=2, value=2256.86).number_format = "$#,##0.00"
            ws.cell(row=r, column=6, value="⚠ Typed in from a calculator. It is correct RIGHT NOW — which is "
                                           "what makes it dangerous. Change the rate and it silently lies.").alignment = WRAP
            ws.cell(row=r, column=6).font = Font(color=WARN)
            r += 2
            for i, h in enumerate(["Month", "Interest", "Principal", "Balance"], start=1):
                ws.cell(row=r, column=i, value=h)
            header_row(ws, r, 4)
            t = r + 1
            for i in range(12):
                row = t + i
                ws.cell(row=row, column=1, value=i + 1)
                prev = "12000" if i == 0 else f"D{row-1}"
                ws.cell(row=row, column=2, value=f"={prev}*0.07/12").number_format = "$#,##0.00"
                ws.cell(row=row, column=3, value=f"=$B${pay_row}-B{row}").number_format = "$#,##0.00"
                ws.cell(row=row, column=4, value=f"={prev}-C{row}").number_format = "$#,##0.00"
            ws.cell(row=t + 13, column=1,
                    value="⚠ Only 12 of 60 months built — the schedule stops before the answer.").alignment = WRAP
            ws.cell(row=t + 13, column=1).font = Font(color=WARN)

        else:
            ws.cell(row=r, column=1, value="car loan stuff").font = H2_FONT
            r += 2
            ws.cell(row=r, column=1, value="monthly payment")
            ws.cell(row=r, column=2, value=237.6).number_format = "$#,##0.00"
            r += 1
            ws.cell(row=r, column=1, value="total")
            ws.cell(row=r, column=2, value=14256)
            r += 1
            ws.cell(row=r, column=1, value="interest")
            ws.cell(row=r, column=2, value=2256)
            r += 2
            for msg in [
                "✕ Every number is typed. There is not one formula in this sheet.",
                "✕ Nothing states the rate or the term, so the answer cannot be checked.",
                "✕ No amortisation schedule, no chart, no conclusion.",
                "✕ Change any assumption and this sheet says exactly the same thing — which is how you know "
                "it is not a model.",
            ]:
                c = ws.cell(row=r, column=1, value=msg)
                c.font = Font(color=BAD)
                c.alignment = WRAP
                r += 1

    # ── What Changed ──
    ws = wb.create_sheet("What Changed")
    set_widths(ws, [16, 52, 52])
    r = title_block(ws, "What Changed Between Levels",
                    "The specific difference — not a vibe. Each row is one concrete move up.")
    for i, h in enumerate(["Step", "What was wrong below", "What fixes it"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 3)
    r += 1
    steps = [
        ("1 → 2",
         "Every number typed in by hand; no formulas at all. The sheet cannot be checked or reused.",
         "Compute at least the payment with a real formula (PMT) and start an amortisation schedule."),
        ("2 → 3",
         "Rate and term buried inside formulas; total interest typed from a calculator; schedule stops at "
         "month 12.",
         "Lift every assumption into its own labelled input cell, point the formulas at those cells, and "
         "build all 60 months so the schedule actually reaches zero."),
        ("3 → 4",
         "Works, but only models the amount financed — no vehicle price or down payment, no protection "
         "against nonsense inputs, chart axes unlabelled.",
         "Add price and down payment as inputs with amount financed derived from them; guard the edge case "
         "with IF(); label both axes and title the chart with the question it answers."),
    ]
    for step, wrong, fix in steps:
        ws.cell(row=r, column=1, value=step).font = BOLD
        ws.cell(row=r, column=2, value=wrong).alignment = WRAP
        ws.cell(row=r, column=3, value=fix).alignment = WRAP
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="The test").font = BOLD
    ws.cell(row=r, column=2,
            value="Change one input. If the answer does not move, you are at Level 1 no matter how it looks.").alignment = WRAP

    path = os.path.join(RES, "csm-u2-exemplars.xlsx")
    wb.save(path)
    return path


def main():
    os.makedirs(RES, exist_ok=True)
    for fn in (build_p2_interest_practice, build_u2_exemplars,
               build_p3_budget_practice,
               build_p4_stocks_practice, build_u4_exemplars):
        p = fn()
        print(f"wrote {os.path.relpath(p, ROOT)}")




# ═══════════════════════════════════════════════════════════════════
# Cornerstone III — young professional in Miami
# ═══════════════════════════════════════════════════════════════════

# Approximate 1-bedroom medians. These are MODELLING VALUES, not quotes —
# drill 2 asks students to verify one against a live listing.
MIAMI_NEIGHBORHOODS = [
    ("Brickell", 3000, 10, "High-rise downtown core; walkable, priciest"),
    ("Downtown", 2700, 10, "Metromover access; busy, mixed"),
    ("Edgewater", 2900, 15, "Bayfront, newer towers"),
    ("Wynwood", 2800, 15, "Arts district; converted warehouses"),
    ("Design District", 3200, 15, "Luxury retail; small rental stock"),
    ("Midtown", 2850, 15, "Between Wynwood and Design District"),
    ("Coral Gables", 2900, 25, "Tree-lined, quiet, good commute south"),
    ("Coconut Grove", 3100, 25, "Waterfront, village feel"),
    ("South Miami", 2400, 30, "Suburban, near the Metrorail"),
    ("Little Havana", 1900, 15, "Historic Cuban core; older stock"),
    ("Allapattah", 1800, 15, "Rapidly changing; near the hospital district"),
    ("Little Haiti", 1750, 20, "Historic; rising quickly"),
    ("Hialeah", 1700, 30, "Largest inventory; longest drive"),
    ("North Miami", 1800, 30, "Near FIU Biscayne Bay"),
    ("Kendall", 2100, 40, "Suburban southwest; car required"),
    ("Doral", 2500, 30, "Near the airport; newer construction"),
    ("Miami Beach", 2600, 35, "Across the causeway; parking is a cost"),
]

MIAMI_JOBS = [
    ("Junior software developer", 72000),
    ("Marketing coordinator", 48000),
    ("Registered nurse", 75000),
    ("Civil engineer (EIT)", 68000),
    ("Paralegal", 52000),
    ("Financial analyst", 70000),
    ("Teacher, first year", 47500),
    ("Graphic designer", 50000),
    ("Mechanical engineer", 71000),
    ("Sales representative", 55000),
    ("Lab technician", 46000),
    ("Architectural designer", 58000),
]

MIAMI_CURVEBALLS = [
    ("Hurricane deductible after a named storm", 2500, "one-time"),
    ("Car repair — pothole on the 826", 900, "one-time"),
    ("Rent hike at renewal (+8%)", 0, "recurring: rent x 1.08"),
    ("Hours cut 15% for two months", 0, "recurring: income x 0.85"),
    ("AC compressor dies in August", 1200, "one-time"),
    ("Health premium increase", 85, "recurring: monthly"),
    ("Towed and impounded during King Tide flooding", 600, "one-time"),
    ("Performance bonus", -2000, "one-time (a gain)"),
]


def build_p3_budget_practice():
    wb = Workbook()
    read_me(wb, "Cornerstone III Practice — Your Miami Month", [
        ("The setup",
         "You have graduated and taken a job in Miami. The randomiser deals you a role, a salary, a "
         "neighborhood, and a student-loan balance. Everybody's hand is different, so there is nothing to copy."),
        ("Roll once, then lock it",
         "RANDBETWEEN re-rolls every time the sheet recalculates. Roll your hand, then select those cells, copy, "
         "and Paste Special > Values to freeze them. That is your life for this cornerstone."),
        ("Florida has no state income tax",
         "That is a real advantage and drill 3 makes you put a number on it. Federal income tax and FICA still "
         "apply — take-home is not salary divided by twelve."),
        ("Rents are approximate",
         "The neighborhood table holds modelling values, not quotes. Drill 2 asks you to check one against a "
         "real listing and report the gap. Models built on unchecked numbers are how people get hurt."),
        ("The point",
         "A budget that only works in the expected month is not a budget. You will break yours on purpose."),
    ])

    # ── Reference tables ──
    ws = wb.create_sheet("Reference Tables")
    set_widths(ws, [30, 16, 16, 46])
    r = title_block(ws, "Reference Tables", "Look values up from here. Never type them into your model.")
    ws.cell(row=r, column=1, value="NEIGHBORHOODS").font = H2_FONT
    r += 1
    for i, h in enumerate(["Neighborhood", "Median 1BR rent", "Commute (min)", "Notes"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 4)
    nb_first = r + 1
    for name, rent, commute, note in MIAMI_NEIGHBORHOODS:
        r += 1
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=rent).number_format = "$#,##0"
        ws.cell(row=r, column=3, value=commute)
        ws.cell(row=r, column=4, value=note).alignment = WRAP
    nb_last = r
    r += 2
    ws.cell(row=r, column=1, value="ROLES").font = H2_FONT
    r += 1
    for i, h in enumerate(["Job title", "Gross salary"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 2)
    job_first = r + 1
    for title, sal in MIAMI_JOBS:
        r += 1
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=sal).number_format = "$#,##0"
    job_last = r
    r += 2
    ws.cell(row=r, column=1, value="CURVEBALLS").font = H2_FONT
    r += 1
    for i, h in enumerate(["Event", "Amount", "How it hits"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 3)
    cb_first = r + 1
    for ev, amt, how in MIAMI_CURVEBALLS:
        r += 1
        ws.cell(row=r, column=1, value=ev)
        ws.cell(row=r, column=2, value=amt).number_format = "$#,##0"
        ws.cell(row=r, column=3, value=how)
    cb_last = r

    # ── Roll your life ──
    ws = wb.create_sheet("1 Roll Your Life")
    set_widths(ws, [26, 30, 16, 52])
    r = title_block(ws, "Drill 1 — Roll your life",
                    "Press F9 to re-roll. When you like your hand — or accept the one you got — copy these "
                    "cells and Paste Special > Values to lock them.")
    ref = "'Reference Tables'"
    rows = [
        ("Neighborhood roll", f"=RANDBETWEEN(1,{nb_last - nb_first + 1})", "0"),
        ("Role roll", f"=RANDBETWEEN(1,{job_last - job_first + 1})", "0"),
        ("Student loan balance", "=RANDBETWEEN(0,45)*1000", "$#,##0"),
    ]
    first = r
    for name, f, fmt in rows:
        ws.cell(row=r, column=1, value=name).font = BOLD
        c = ws.cell(row=r, column=2, value=f); c.number_format = fmt; c.fill = BAND_FILL
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="YOUR HAND").font = H2_FONT
    r += 1
    hand = [
        ("Neighborhood", f"=INDEX({ref}!$A${nb_first}:$A${nb_last},$B${first})", "General"),
        ("Median 1BR rent", f"=INDEX({ref}!$B${nb_first}:$B${nb_last},$B${first})", "$#,##0"),
        ("Commute (min each way)", f"=INDEX({ref}!$C${nb_first}:$C${nb_last},$B${first})", "0"),
        ("Job title", f"=INDEX({ref}!$A${job_first}:$A${job_last},$B${first+1})", "General"),
        ("Gross salary", f"=INDEX({ref}!$B${job_first}:$B${job_last},$B${first+1})", "$#,##0"),
        ("Student loan balance", f"=$B${first+2}", "$#,##0"),
    ]
    hand_rows = {}
    for name, f, fmt in hand:
        ws.cell(row=r, column=1, value=name).font = BOLD
        c = ws.cell(row=r, column=2, value=f)
        if fmt != "General":
            c.number_format = fmt
        hand_rows[name] = r
        r += 1
    ROLL_RENT = f"'1 Roll Your Life'!B{hand_rows['Median 1BR rent']}"
    ROLL_SALARY = f"'1 Roll Your Life'!B{hand_rows['Gross salary']}"
    ws.cell(row=first, column=4,
            value="INDEX pulls the row the roll landed on. Read the formula — this is the same lookup move "
                  "from Cornerstone I, doing real work.").alignment = WRAP

    globals()["_ROLL_RENT"] = ROLL_RENT
    globals()["_ROLL_SALARY"] = ROLL_SALARY

    # ── Verify the rent ──
    ws = wb.create_sheet("2 Verify The Rent")
    set_widths(ws, [34, 20, 62])
    r = title_block(ws, "Drill 2 — Check the number against reality",
                    "The reference table holds approximate medians. Find one real listing in your neighborhood.")
    for label in ["Your neighborhood", "Table value (median 1BR)", "Real listing you found (URL)",
                  "Real listing rent", "Difference", "Percent difference"]:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2).fill = BAND_FILL
        r += 1
    ws.cell(row=r - 2, column=2, value="=B{}-B{}".format(r - 3, r - 5)).number_format = "$#,##0"
    ws.cell(row=r - 1, column=2, value="=IF(B{}=0,\"\",(B{}-B{})/B{})".format(r - 5, r - 3, r - 5, r - 5)).number_format = "0.0%"
    r += 1
    ws.cell(row=r, column=1,
            value="Which number will you build your budget on, and why? (Two sentences.)").font = BOLD
    r += 1
    ws.cell(row=r, column=1, value="").fill = BAND_FILL

    # ── Take-home pay ──
    ws = wb.create_sheet("3 Take-Home Pay")
    set_widths(ws, [32, 18, 16, 54])
    r = title_block(ws, "Drill 3 — Salary is not take-home",
                    "Build gross to net. Florida has no state income tax; put a dollar value on that.")
    ws.cell(row=r, column=1, value="Gross annual salary").font = BOLD
    ws.cell(row=r, column=2, value=f"={_ROLL_SALARY}").number_format = "$#,##0"
    g = r; r += 1
    steps = [
        ("Standard deduction (single, approx.)", "=15000", "$#,##0"),
        ("Taxable income", f"=MAX(0,B{g}-B{g+1})", "$#,##0"),
        ("Federal income tax (approx. effective)", f"=B{g+2}*0.12", "$#,##0"),
        ("FICA (7.65%)", f"=B{g}*0.0765", "$#,##0"),
        ("Florida state income tax", "=0", "$#,##0"),
        ("Net annual", f"=B{g}-B{g+3}-B{g+4}-B{g+5}", "$#,##0"),
        ("Take-home per month", f"=B{g+6}/12", "$#,##0.00"),
    ]
    step_rows = {}
    for name, f, fmt in steps:
        ws.cell(row=r, column=1, value=name).font = BOLD
        ws.cell(row=r, column=2, value=f).number_format = fmt
        step_rows[name] = r
        r += 1
    globals()["_TAKEHOME"] = f"'3 Take-Home Pay'!B{step_rows['Take-home per month']}"
    ws.cell(row=g + 5, column=4,
            value="Zero — and that is the point. Drill: what would this cell say if you took the same job in "
                  "New York (about 6%) or California (about 8%)? Compute both and write the monthly "
                  "difference below.").alignment = WRAP
    r += 1
    ws.cell(row=r, column=1, value="Monthly value of living in Florida vs New York").font = BOLD
    ws.cell(row=r, column=2, value=f"=B{g+2}*0.06/12").number_format = "$#,##0.00"
    r += 1
    ws.cell(row=r, column=1, value="Monthly value vs California").font = BOLD
    ws.cell(row=r, column=2, value=f"=B{g+2}*0.08/12").number_format = "$#,##0.00"

    # ── The month, baseline and broken ──
    ws = wb.create_sheet("4-5 Baseline vs Curveball")
    set_widths(ws, [30, 16, 16, 16, 50])
    r = title_block(ws, "Drills 4 & 5 — The month, twice",
                    "Classify every line as fixed or variable, then run it again with a curveball.")
    for i, h in enumerate(["Category", "Fixed or variable?", "Baseline", "With curveball"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 4)
    r += 1
    cats = [("Rent", "Fixed", f"={_ROLL_RENT}"), ("Renter's insurance", "Fixed", 25),
            ("Electricity (FPL)", "Variable", 140), ("Internet", "Fixed", 70),
            ("Phone", "Fixed", 55), ("Groceries", "Variable", 400),
            ("Eating out", "Variable", 220), ("Car payment", "Fixed", 380),
            ("Car insurance (Miami rates)", "Fixed", 245), ("Gas / tolls (SunPass)", "Variable", 160),
            ("Parking", "Fixed", 90), ("Student loan payment", "Fixed", 250),
            ("Health insurance", "Fixed", 210), ("Subscriptions", "Variable", 45),
            ("Going out", "Variable", 180)]
    first_cat = r
    for name, kind, val in cats:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=kind)
        c = ws.cell(row=r, column=3, value=val); c.number_format = "$#,##0"
        ws.cell(row=r, column=4, value=f"=C{r}").number_format = "$#,##0"
        r += 1
    last_cat = r - 1
    ws.cell(row=r, column=1, value="TOTAL SPENDING").font = BOLD
    ws.cell(row=r, column=3, value=f"=SUM(C{first_cat}:C{last_cat})").number_format = "$#,##0"
    ws.cell(row=r, column=4, value=f"=SUM(D{first_cat}:D{last_cat})").number_format = "$#,##0"
    tot = r; r += 1
    ws.cell(row=r, column=1, value="Take-home").font = BOLD
    ws.cell(row=r, column=3, value=f"={_TAKEHOME}").number_format = "$#,##0"
    ws.cell(row=r, column=4, value=f"=C{r}").number_format = "$#,##0"
    inc = r; r += 1
    ws.cell(row=r, column=1, value="LEFT OVER").font = BOLD
    ws.cell(row=r, column=3, value=f"=C{inc}-C{tot}").number_format = "$#,##0"
    ws.cell(row=r, column=4, value=f"=D{inc}-D{tot}").number_format = "$#,##0"
    left = r; r += 1
    ws.cell(row=r, column=1, value="SAVINGS RATE").font = BOLD
    ws.cell(row=r, column=3, value=f"=IF(C{inc}=0,\"\",C{left}/C{inc})").number_format = "0.0%"
    ws.cell(row=r, column=4, value=f"=IF(D{inc}=0,\"\",D{left}/D{inc})").number_format = "0.0%"
    r += 2
    ws.cell(row=r, column=1, value="Months until savings hit zero (curveball month repeating)").font = BOLD
    ws.cell(row=r, column=3, value=f"=IF(D{left}>=0,\"never — it holds\",ROUND(1000/-D{left},1))")
    ws.cell(row=r, column=5,
            value="Drill 5: pick a curveball from the Reference Tables and apply it in column D only. "
                  "Assume $1,000 in savings. Does your month hold?").alignment = WRAP

    path = os.path.join(RES, "csm-p3-budget-practice.xlsx")
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════
# Cornerstone IV — the Investopedia $100,000 challenge
# ═══════════════════════════════════════════════════════════════════

VERTICALS = [
    ("Technology", "software, semiconductors, hardware"),
    ("Healthcare", "pharma, devices, insurers"),
    ("Consumer", "retail, food and beverage, apparel"),
    ("Energy & Utilities", "oil and gas, renewables, power"),
    ("Financials", "banks, payments, insurance"),
    ("Industrials & Transport", "manufacturing, logistics, airlines"),
]

THESIS_FIELDS = [
    "Ticker",
    "Vertical",
    "Shares",
    "Entry price",
    "Position value",
    "% of portfolio",
    "WHY this company (the thesis)",
    "What I expect to happen",
    "What would prove me WRONG",
]


def build_p4_stocks_practice():
    wb = Workbook()
    read_me(wb, "Cornerstone IV Practice — Research, Returns & Simulation", [
        ("The challenge",
         "You are handed $100,000 of simulated capital in the Investopedia Stock Simulator. You must hold at "
         "least one position in each of six verticals — you may not put it all in tech."),
        ("The rule that matters",
         "Every trade needs a written thesis BEFORE it is placed: why this company, what you expect, and what "
         "would prove you wrong. A trade with no recorded thesis scores zero regardless of what it returns."),
        ("You are not graded on returns",
         "A well-argued loss outscores a lucky win. Markets are noisy over a semester; reasoning is what we "
         "can actually assess, and it is the thing that transfers."),
        ("Why the model matters",
         "Without a forecast you cannot tell skill from luck. The Monte Carlo work gives you an expected "
         "range, so you can say whether your result was surprising or ordinary."),
    ])

    # ── Trade log ──
    ws = wb.create_sheet("1 Trade Log")
    set_widths(ws, [10, 22, 10, 14, 16, 13, 46, 40, 40])
    r = title_block(ws, "Drill 1 — Record the thesis BEFORE you buy",
                    "Fill the right-hand three columns first. If you cannot write them, you are not ready to "
                    "place the trade.")
    for i, h in enumerate(THESIS_FIELDS, start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, len(THESIS_FIELDS))
    first = r + 1
    for i, (vert, _) in enumerate(VERTICALS):
        row = first + i
        ws.cell(row=row, column=2, value=vert)
        ws.cell(row=row, column=5, value=f"=IF(AND(C{row}<>\"\",D{row}<>\"\"),C{row}*D{row},\"\")").number_format = "$#,##0.00"
        ws.cell(row=row, column=6, value=f"=IF(E{row}=\"\",\"\",E{row}/$B${first+len(VERTICALS)+3})").number_format = "0.0%"
        for c in (1, 3, 4, 7, 8, 9):
            ws.cell(row=row, column=c).fill = BAND_FILL
        for c in (7, 8, 9):
            ws.cell(row=row, column=c).alignment = WRAP
        ws.row_dimensions[row].height = 42
    r = first + len(VERTICALS) + 1
    ws.cell(row=r, column=1, value="Invested").font = BOLD
    ws.cell(row=r, column=5, value=f"=SUM(E{first}:E{first+len(VERTICALS)-1})").number_format = "$#,##0.00"
    r += 1
    ws.cell(row=r, column=1, value="Cash remaining").font = BOLD
    ws.cell(row=r, column=5, value=f"=$B${first+len(VERTICALS)+3}-E{r-1}").number_format = "$#,##0.00"
    r += 1
    ws.cell(row=r, column=1, value="Starting capital").font = BOLD
    c = ws.cell(row=r, column=2, value=100000); c.number_format = "$#,##0"; c.fill = BAND_FILL

    # ── Returns and volatility ──
    ws = wb.create_sheet("2-4 Returns & Volatility")
    set_widths(ws, [12, 14, 16, 16, 16, 50])
    r = title_block(ws, "Drills 2–4 — What the price series tells you",
                    "Paste 60 trading days of closing prices for ONE holding, then build the three measures.")
    for i, h in enumerate(["Day", "Close", "Daily return", "5-day MA", ""], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 4)
    t = r + 1
    for i in range(60):
        row = t + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2).fill = BAND_FILL
        if i > 0:
            ws.cell(row=row, column=3, value=f"=IF(OR(B{row}=\"\",B{row-1}=\"\"),\"\",(B{row}-B{row-1})/B{row-1})").number_format = "0.00%"
        if i >= 4:
            ws.cell(row=row, column=4, value=f"=IF(COUNT(B{row-4}:B{row})<5,\"\",AVERAGE(B{row-4}:B{row}))").number_format = "$#,##0.00"
    r = t + 61
    for name, f, fmt in [
        ("Mean daily return", f"=IFERROR(AVERAGE(C{t+1}:C{t+59}),\"\")", "0.000%"),
        ("Daily volatility (st dev)", f"=IFERROR(STDEV(C{t+1}:C{t+59}),\"\")", "0.000%"),
        ("Annualised volatility", f"=IFERROR(B{r+1}*SQRT(252),\"\")", "0.0%"),
    ]:
        ws.cell(row=r, column=1, value=name).font = BOLD
        ws.cell(row=r, column=2, value=f).number_format = fmt
        r += 1
    ws.cell(row=t, column=6,
            value="Drill 4: run this sheet for a utility and again for a growth stock. Write one sentence "
                  "on what the volatility difference means for someone holding each.").alignment = WRAP

    # ── Monte Carlo ──
    ws = wb.create_sheet("5-6 Monte Carlo")
    set_widths(ws, [26, 16, 16, 16, 16, 50])
    r = title_block(ws, "Drills 5 & 6 — A thousand futures",
                    "You cannot forecast one outcome. Forecast the distribution, then read it honestly.")
    ws.cell(row=r, column=1, value="INPUTS").font = H2_FONT
    r += 1
    base = r
    for name, val, fmt in [("Starting value", 100000, "$#,##0"),
                           ("Expected annual return", 0.08, "0.0%"),
                           ("Annual volatility", 0.18, "0.0%"),
                           ("Years", 1, "0")]:
        ws.cell(row=r, column=1, value=name).font = BOLD
        c = ws.cell(row=r, column=2, value=val); c.number_format = fmt; c.fill = BAND_FILL
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="200 simulated outcomes").font = H2_FONT
    r += 1
    ws.cell(row=r, column=1, value="Trial")
    ws.cell(row=r, column=2, value="Ending value")
    header_row(ws, r, 2)
    sim = r + 1
    for i in range(200):
        row = sim + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2,
                value=f"=$B${base}*EXP(($B${base+1}-0.5*$B${base+2}^2)*$B${base+3}"
                      f"+$B${base+2}*SQRT($B${base+3})*NORM.S.INV(RAND()))").number_format = "$#,##0"
    r = sim + 201
    for name, f in [("10th percentile", f"=PERCENTILE(B{sim}:B{sim+199},0.1)"),
                    ("50th percentile (median)", f"=PERCENTILE(B{sim}:B{sim+199},0.5)"),
                    ("90th percentile", f"=PERCENTILE(B{sim}:B{sim+199},0.9)"),
                    ("Chance of losing money", f"=COUNTIF(B{sim}:B{sim+199},\"<\"&$B${base})/200")]:
        ws.cell(row=r, column=1, value=name).font = BOLD
        cell = ws.cell(row=r, column=2, value=f)
        cell.number_format = "0.0%" if "Chance" in name else "$#,##0"
        r += 1
    ws.cell(row=sim, column=4,
            value="Drill 6: quote all three percentiles in your write-up. Reporting only the median hides the "
                  "risk, and hiding risk is the most common dishonesty in finance.").alignment = WRAP

    path = os.path.join(RES, "csm-p4-stocks-practice.xlsx")
    wb.save(path)
    return path


def build_u4_exemplars():
    """Four levels of the Investopedia deliverable. The differentiator here is
    the quality of the recorded thesis, not the return — so each level shows
    the SAME trade argued at a different standard."""
    wb = Workbook()
    read_me(wb, "Cornerstone IV — What Each Level Looks Like", [
        ("What this is",
         "The same holding — one position in Healthcare — written up four times. Every level bought the same "
         "stock at the same price. What separates them is entirely the reasoning."),
        ("Why it looks like this",
         "You are not graded on whether the position went up. You are graded on whether you could say, before "
         "you bought it, why you were buying and what would prove you wrong."),
        ("The test for a real thesis",
         "It must be falsifiable. 'It will go up' cannot be wrong in any specific way, so it is not a thesis. "
         "'Margins compress if the patent expires in Q3' can be checked."),
        ("Read the Falsifiable? column",
         "That single column is the fastest predictor of the grade."),
    ])

    ROWS = [
        (4, "Advanced", GOOD,
         "JNJ", "Healthcare", 120, 162.40,
         "Medical devices and consumer health are defensive: demand does not fall when the economy does. "
         "Trading at ~15x forward earnings against a 5-year average of ~17x, and the talc litigation that "
         "drove the discount was settled in Q1. I want one holding that does not move with the tech names.",
         "Slow appreciation, roughly market return, with materially lower drawdown than my tech position. "
         "I expect it to hold up on the days the Nasdaq falls hard.",
         "If it falls more than 8% on a day the S&P falls 2%, my defensive premise is wrong and I should "
         "exit — the correlation I bought is not there.",
         "Yes — a specific, checkable condition with an action attached."),
        (3, "Proficient", GOOD,
         "JNJ", "Healthcare", 120, 162.40,
         "Healthcare is a defensive sector and JNJ is one of the largest, most stable companies in it. It pays "
         "a dividend and is less volatile than my other holdings.",
         "Steady, modest growth. Less movement than the rest of the portfolio.",
         "If it drops sharply along with everything else, then it is not actually defensive.",
         "Yes — checkable, though vaguer about the threshold and the action."),
        (2, "Developing", WARN,
         "JNJ", "Healthcare", 120, 162.40,
         "It is a big healthcare company and healthcare is a required vertical. It seems safe and everyone "
         "knows the brand.",
         "It should go up over time.",
         "If it goes down.",
         "No — 'if it goes down' is not a falsifiable premise, it is a restatement of loss."),
        (1, "Beginning", BAD,
         "JNJ", "Healthcare", 120, 162.40,
         "Needed something for healthcare.",
         "", "",
         "No thesis recorded at all. Scores zero regardless of return."),
    ]

    for level, name, colour, tk, vert, sh, px, why, expect, wrong, verdict in ROWS:
        ws = wb.create_sheet(f"Level {level} — {name}")
        set_widths(ws, [24, 62, 62])
        ws["A1"] = f"LEVEL {level} — {name}"
        ws["A1"].font = Font(bold=True, size=15, color=colour)
        ws["A2"] = "Same stock, same price, same day. Only the reasoning differs."
        ws["A2"].font = NOTE_FONT
        r = 4
        facts = [("Ticker", tk), ("Vertical", vert), ("Shares", sh),
                 ("Entry price", px), ("Position value", f"={sh}*{px}"),
                 ("% of $100,000 portfolio", f"={sh}*{px}/100000")]
        for lbl, val in facts:
            ws.cell(row=r, column=1, value=lbl).font = BOLD
            c = ws.cell(row=r, column=2, value=val)
            if lbl in ("Entry price", "Position value"):
                c.number_format = "$#,##0.00"
            if lbl.startswith("%"):
                c.number_format = "0.0%"
            r += 1
        r += 1
        for lbl, body in [("WHY this company", why),
                          ("What I expect", expect or "— left blank —"),
                          ("What would prove me WRONG", wrong or "— left blank —")]:
            ws.cell(row=r, column=1, value=lbl).font = BOLD
            c = ws.cell(row=r, column=2, value=body)
            c.alignment = WRAP
            ws.row_dimensions[r].height = max(30, 13 * (1 + len(body) // 60))
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="Falsifiable?").font = BOLD
        c = ws.cell(row=r, column=2, value=verdict)
        c.alignment = WRAP
        c.font = Font(color=colour, bold=True)

    ws = wb.create_sheet("What Changed")
    set_widths(ws, [16, 54, 54])
    r = title_block(ws, "What Changed Between Levels",
                    "Every level bought the same stock. Here is what actually moved the grade.")
    for i, h in enumerate(["Step", "What was missing below", "What fixes it"], start=1):
        ws.cell(row=r, column=i, value=h)
    header_row(ws, r, 3)
    r += 1
    for step, missing, fix in [
        ("1 → 2", "No thesis at all — the position exists only because a vertical had to be filled.",
         "Write something before you buy, even if it is thin. A recorded reason is the minimum bar."),
        ("2 → 3", "'If it goes down' is not falsifiable — it restates loss instead of naming a condition. "
                  "No evidence, no numbers.",
         "Name a property of the company you are betting on (defensive, dividend, low volatility) and a "
         "condition that would show the bet was wrong."),
        ("3 → 4", "The claim is right but unquantified: no valuation evidence, no threshold, no action if the "
                  "thesis breaks.",
         "Add the number that made it cheap (15x vs a 17x average), the catalyst that explains the discount, "
         "a measurable disproof threshold (−8% on a −2% day), and what you will DO when it triggers."),
    ]:
        ws.cell(row=r, column=1, value=step).font = BOLD
        ws.cell(row=r, column=2, value=missing).alignment = WRAP
        ws.cell(row=r, column=3, value=fix).alignment = WRAP
        ws.row_dimensions[r].height = 52
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="The test").font = BOLD
    ws.cell(row=r, column=2,
            value="Can your 'what would prove me wrong' actually happen and be observed? If not, rewrite it "
                  "before you place the trade.").alignment = WRAP
    r += 2
    ws.cell(row=r, column=1, value="Note on the numbers").font = BOLD
    ws.cell(row=r, column=2,
            value="The ticker, price, and valuation figures above are ILLUSTRATIVE — written to show what a "
                  "good argument looks like, not as investment guidance or current market data. Verify "
                  "everything against Investopedia before you trade.").alignment = WRAP

    path = os.path.join(RES, "csm-u4-exemplars.xlsx")
    wb.save(path)
    return path

if __name__ == "__main__":
    main()
